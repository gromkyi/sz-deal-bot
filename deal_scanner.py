#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
STALZONE deal scanner.
Почасово проходит по watchlist.csv, сравнивает минимальную цену покупки (лоты)
с медианой реальных сделок, считает ЧИСТУЮ маржу после комиссии аукциона,
и шлёт в Телеграм выгодные лоты, сгруппированные по ценовым блокам.

Все пороги — в config.json (обратная прогрессия маржи, бюджет, блоки и т.д.).
Учётки/токены — из переменных окружения (GitHub Actions secrets) или локальных файлов:
  STALZONE_CLIENT_ID, STALZONE_CLIENT_SECRET, TG_BOT_TOKEN, TG_CHAT_ID
Только стандартная библиотека Python 3.
"""
import csv, os, sys, json, time, statistics, datetime, html
import urllib.request, urllib.error, urllib.parse

HERE = os.path.dirname(os.path.abspath(__file__))
API_BASE = os.environ.get("STALZONE_API_BASE", "https://eapi.stalcraft.net")
REGION   = os.environ.get("STALZONE_REGION", "RU")

DATA_DIR = os.environ.get("DATA_DIR", HERE)   # на Railway указать на том (volume), чтобы антиспам/лог не сбрасывались
WATCH = os.path.join(HERE, "watchlist.csv")
STATE = os.path.join(DATA_DIR, "alerts_state.json")
DEALS_LOG = os.path.join(DATA_DIR, "deals_log.csv")
CONFIG = os.path.join(HERE, "config.json")
WISHLIST = os.path.join(HERE, "wishlist.csv")

DEFAULTS = {
    "budget": {"min_buy": 5000, "max_buy": 1000000},
    "min_liquidity_per_day": 5,
    "auction_fee": 0.05,
    "margin_tiers": [
        {"max_price": 10000, "min_margin_pct": 25},
        {"max_price": 100000, "min_margin_pct": 20},
        {"max_price": 1000000, "min_margin_pct": 15},
        {"max_price": None, "min_margin_pct": 10},
    ],
    "display_blocks": [
        {"label": "5–50k", "max": 50000},
        {"label": "50–100k", "max": 100000},
        {"label": "100–500k", "max": 500000},
        {"label": "500k–1M", "max": 1000000},
        {"label": ">1M", "max": None},
    ],
    "alert_cooldown_hours": 6,
    "max_items_per_block": 10,
    "sort_within_block": "profit",
}
LOTS_LIMIT, HIST_LIMIT = 200, 100
SLEEP_SEC = float(os.environ.get("SLEEP_SEC", "0.25"))

def load_config():
    cfg = dict(DEFAULTS)
    if os.path.exists(CONFIG):
        try:
            user = json.load(open(CONFIG, encoding="utf-8"))
            for k, v in user.items():
                if not k.startswith("_"): cfg[k] = v
        except Exception as e:
            print("config.json ошибка, использую дефолты:", e)
    return cfg

def cred(env_key, file_name, json_key):
    v = os.environ.get(env_key)
    if v: return v.strip()
    p = os.path.join(HERE, file_name)
    if os.path.exists(p):
        return str(json.load(open(p, encoding="utf-8")).get(json_key, "")).strip()
    return ""

def headers():
    cid = cred("STALZONE_CLIENT_ID", "api_credentials.json", "client_id")
    sec = cred("STALZONE_CLIENT_SECRET", "api_credentials.json", "client_secret")
    if not cid or not sec:
        sys.exit("Нет client_id/client_secret (env или api_credentials.json)")
    return {"Client-Id": cid, "Client-Secret": sec, "Accept": "application/json",
            "User-Agent": "StalzoneDealScanner/1.1"}

def api(path, H):
    req = urllib.request.Request(f"{API_BASE}/{REGION}/{path}", headers=H)
    with urllib.request.urlopen(req, timeout=25) as r:
        return json.loads(r.read().decode("utf-8"))

def fnum(x):
    try: return float(x)
    except (TypeError, ValueError): return None

def cheapest_ask(data):
    lots = (data.get("lots") if isinstance(data, dict) else data) or []
    best = None
    for lot in lots:
        amt = fnum(lot.get("amount")) or 1
        p = fnum(lot.get("buyoutPrice"))
        if p is None or p <= 0: continue
        u = p/amt if amt else p
        if best is None or u < best: best = u
    return best

def unit_asks(data, want_key=None):
    """Отсортированный список цен за 1 шт по текущим лотам с выкупом.
    want_key (для артефактов) — учитывать только лоты этого качества (см. quality_key)."""
    lots = (data.get("lots") if isinstance(data, dict) else data) or []
    out = []
    for lot in lots:
        amt = fnum(lot.get("amount")) or 1
        p = fnum(lot.get("buyoutPrice"))
        if p is None or p <= 0: continue
        if want_key is not None and quality_key(lot.get("additional")) != want_key: continue
        out.append(p/amt if amt else p)
    out.sort()
    return out

def market(data):
    prices = (data.get("prices") if isinstance(data, dict) else data) or []
    units, times = [], []
    for s in prices:
        p = fnum(s.get("price"))
        if p is None: continue
        amt = fnum(s.get("amount")) or 1
        units.append(p/amt if amt else p)   # price — за весь стак, делим на количество
        t = s.get("time")
        if t:
            try: times.append(datetime.datetime.fromisoformat(t.replace("Z", "+00:00")))
            except Exception: pass
    med = round(statistics.median(units), 2) if units else None
    liq = None
    if len(times) >= 2:
        span = (max(times)-min(times)).total_seconds()/86400
        liq = (len(times)-1)/span if span > 0 else 999
    return med, liq

def threshold_for(price, tiers):
    for t in tiers:
        mx = t.get("max_price")
        if mx is None or price < mx:
            return float(t.get("min_margin_pct", 0))
    return float(tiers[-1].get("min_margin_pct", 0))

def block_for(price, blocks):
    for b in blocks:
        mx = b.get("max")
        if mx is None or price < mx:
            return b.get("label", "")
    return blocks[-1].get("label", "")

def sp(n):  # число с пробелами-разделителями тысяч
    return f"{int(round(n)):,}".replace(",", " ")

def section(category):
    c = (category or "").split("/")[0]
    return {"artefact":"artefacts","weapon":"weapon","armor":"armor","attachment":"attachments",
            "weapon_modules":"weapon-modules","container":"containers","backpack":"backpacks",
            "device":"devices"}.get(c, "consumables")

def _tg_send_one(url, chat, text, parse_mode="HTML"):
    data = {"chat_id": chat, "text": text, "disable_web_page_preview": "true"}
    if parse_mode: data["parse_mode"] = parse_mode
    payload = urllib.parse.urlencode(data).encode()
    with urllib.request.urlopen(urllib.request.Request(url, data=payload), timeout=20) as r:
        r.read()

def send_telegram(text):
    token = cred("TG_BOT_TOKEN", "telegram_config.json", "bot_token")
    chat = cred("TG_CHAT_ID", "telegram_config.json", "chat_id")
    if not token or not chat:
        print("Telegram не настроен — печатаю в консоль:\n" + text); return
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    # Телеграм-лимит 4096 символов: режем по строкам на части <=3800
    chunks, cur = [], ""
    for line in text.split("\n"):
        if len(cur) + len(line) + 1 > 3800:
            if cur: chunks.append(cur)
            cur = line
        else:
            cur = (cur + "\n" + line) if cur else line
    if cur: chunks.append(cur)
    for ch in chunks:
        try:
            _tg_send_one(url, chat, ch)
        except urllib.error.HTTPError as e:
            body = ""
            try: body = e.read().decode("utf-8", "replace")
            except Exception: pass
            print("Telegram ошибка:", e.code, body)
            # запасной вариант — без HTML-разметки (на случай проблем с парсингом)
            import re as _re
            try:
                _tg_send_one(url, chat, _re.sub(r"<[^>]+>", "", ch), parse_mode=None)
                print("  -> отправлено без разметки")
            except Exception as e2:
                print("  -> и без разметки не вышло:", e2)
        except Exception as e:
            print("Ошибка отправки в Telegram:", e)
        time.sleep(0.4)

def quality_key(add):
    add = add or {}
    bp = tuple(sorted(add.get("bonus_properties") or []))
    return (add.get("qlt"), add.get("ptn"), bp)

RARITY = {0: ("⚪", "обычный"), 1: ("🟢", "необычный"), 2: ("🟣", "особый"),
          3: ("🌸", "редкий"), 4: ("🔴", "исключительный"), 5: ("🟡", "легендарный")}

def quality_label(add):
    add = add or {}
    parts = []
    q = add.get("qlt")
    if q is not None:
        if q in RARITY:
            em, nm = RARITY[q]; parts.append(f"{em} {nm}")
        else:
            parts.append(f"кач.{q}")
    if add.get("ptn") is not None: parts.append(f"потенц.{add['ptn']}")
    bp = add.get("bonus_properties") or []
    if bp: parts.append(f"бонусов {len(bp)}")
    return ", ".join(parts) if parts else "базовый"

def find_quality_deal(iid, H, cfg, fee, minliq, bmin, bmax):
    """Артефакты: в каждом качестве берёт самый дешёвый лот и сравнивает со ВТОРОЙ мин. ценой
    того же качества (реалистичная перепродажа). Медиана/ликвидность — контекст. Лучший дил или None."""
    lots = api(f"auction/{iid}/lots?limit=200&additional=true", H)
    time.sleep(SLEEP_SEC)
    hist = api(f"auction/{iid}/history?limit=200&additional=true", H)
    min_n = int(cfg.get("min_sales_per_quality", 3))
    # история по качеству -> медиана (контекст) + ликвидность
    sales, times_ = {}, {}
    for s in (hist.get("prices") or []):
        p = fnum(s.get("price")); amt = fnum(s.get("amount")) or 1
        if p is None: continue
        k = quality_key(s.get("additional"))
        sales.setdefault(k, []).append(p/amt if amt else p)
        t = s.get("time")
        if t:
            try: times_.setdefault(k, []).append(datetime.datetime.fromisoformat(t.replace("Z", "+00:00")))
            except Exception: pass
    # текущие лоты по качеству
    qlots = {}
    for lot in (lots.get("lots") or []):
        p = fnum(lot.get("buyoutPrice")); amt = fnum(lot.get("amount")) or 1
        if p is None or p <= 0: continue
        k = quality_key(lot.get("additional"))
        qlots.setdefault(k, []).append((p/amt if amt else p, lot.get("additional")))
    best = None
    for k, arr in qlots.items():
        arr.sort(key=lambda x: x[0])
        if len(arr) < 2: continue                 # нет второй цены — не с чем сравнивать
        ask = arr[0][0]; ref = arr[1][0]          # 1-я и 2-я мин. цена этого качества
        if not (bmin <= ask <= bmax): continue
        ks = sales.get(k, [])
        med = statistics.median(ks) if len(ks) >= min_n else None
        tk = times_.get(k, []); liq = None
        if len(tk) >= 2:
            span = (max(tk)-min(tk)).total_seconds()/86400
            liq = (len(tk)-1)/span if span > 0 else 999
        if liq is not None and liq < minliq: continue
        net = (ref*(1-fee) - ask)/ask*100
        if net < threshold_for(ask, cfg["margin_tiers"]): continue
        below = sum(1 for pr, _ in arr if med is not None and pr < med)
        cand = {"ask": ask, "ref": ref, "med": (round(med) if med is not None else None),
                "net": net, "profit": ref*(1-fee)-ask, "liq": liq,
                "qlabel": quality_label(arr[0][1]), "below": below}
        if best is None or cand["net"] > best["net"]: best = cand
    return best

def load_wishlist():
    """Список покупок: wishlist.xlsx (приоритет) или wishlist.csv.
    Возвращает кортежи (id, name, max_price, discount_pct)."""
    xlsx = os.path.join(HERE, "wishlist.xlsx")
    if os.path.exists(xlsx):
        try:
            from openpyxl import load_workbook
            ws = load_workbook(xlsx, data_only=True).active
            hdr = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[1]]
            def col(*names):
                for i, h in enumerate(hdr):
                    if any(n in h for n in names): return i
                return None
            ic = col("id"); nc = col("предмет", "name", "назв")
            pc = col("макс"); dc = col("скидк", "медиан")
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if ic is None or ic >= len(row) or not row[ic]: continue
                iid = str(row[ic]).strip()
                name = str(row[nc]).strip() if (nc is not None and nc < len(row) and row[nc]) else iid
                mp = row[pc] if (pc is not None and pc < len(row)) else None
                dp = row[dc] if (dc is not None and dc < len(row)) else None
                rows.append((iid, name, mp, dp))
            return rows
        except Exception as e:
            print("wishlist.xlsx ошибка чтения:", e)
    rows = []
    if os.path.exists(WISHLIST):
        for r in csv.DictReader(open(WISHLIST, encoding="utf-8-sig")):
            rows.append(((r.get("id") or "").strip(), r.get("name") or "", r.get("max_price"), r.get("discount_pct")))
    return rows

def main():
    cfg = load_config()
    if not os.path.exists(WATCH):
        sys.exit("Нет watchlist.csv — сначала собери его (build_watchlist.py)")
    H = headers()
    items = list(csv.DictReader(open(WATCH, encoding="utf-8-sig")))
    if "--limit" in sys.argv:   # быстрый тест: проверить только первые N предметов
        try: items = items[:int(sys.argv[sys.argv.index("--limit")+1])]
        except (ValueError, IndexError): pass
    state = json.load(open(STATE, encoding="utf-8")) if os.path.exists(STATE) else {}
    now = datetime.datetime.now(datetime.timezone.utc)

    bmin = float(cfg["budget"]["min_buy"]); bmax = float(cfg["budget"]["max_buy"])
    fee = float(cfg["auction_fee"]); minliq = float(cfg["min_liquidity_per_day"])
    cooldown = float(cfg["alert_cooldown_hours"])*3600
    excl_cats = [str(x) for x in cfg.get("exclude_categories", [])]
    qual_cats = [str(x) for x in cfg.get("quality_categories", [])]
    deal_min = float(cfg.get("deal_min_price", 0) or 0)

    deals, checked = [], 0
    for r in items:
        iid, name, cat = r["id"], r.get("name", r["id"]), r.get("category", "")
        if excl_cats and any(cat.startswith(x) for x in excl_cats): continue
        if deal_min and (fnum(r.get("median_price")) or 0) < deal_min: continue

        # артефакты и т.п. — сравнение по качеству
        if qual_cats and any(cat.startswith(x) for x in qual_cats):
            try:
                res = find_quality_deal(iid, H, cfg, fee, minliq, bmin, bmax); time.sleep(SLEEP_SEC)
            except urllib.error.HTTPError as e:
                if e.code == 429: time.sleep(30)
                continue
            except Exception:
                continue
            checked += 1
            if not res: continue
            qkey = "q:" + iid + ":" + res["qlabel"]
            st = state.get(qkey)
            if st:
                try: last = datetime.datetime.fromisoformat(st["ts"])
                except Exception: last = None
                if last and (now-last).total_seconds() < cooldown and res["ask"] >= st.get("ask", 0)*0.95:
                    continue
            deals.append({"id":iid,"name":name,"cat":cat,"ask":res["ask"],"med":res.get("med"),"ref":res["ref"],
                          "net":round(res["net"],1),"profit":round(res["profit"]),
                          "liq":round(res["liq"],1) if res["liq"] else "n/a",
                          "block":block_for(res["ask"], cfg["display_blocks"]), "ql":res["qlabel"],
                          "below":res.get("below")})
            state[qkey] = {"ask": res["ask"], "ts": now.isoformat()}
            continue

        try:
            asks = unit_asks(api(f"auction/{iid}/lots?limit={LOTS_LIMIT}", H)); time.sleep(SLEEP_SEC)
            med, liq = market(api(f"auction/{iid}/history?limit={HIST_LIMIT}", H)); time.sleep(SLEEP_SEC)
        except urllib.error.HTTPError as e:
            if e.code == 429: time.sleep(30)
            continue
        except Exception:
            continue
        checked += 1
        ask = asks[0] if asks else None
        ref = asks[1] if len(asks) > 1 else None   # 2-я мин. цена рынка = реалистичная перепродажа
        if ask is None or ref is None: continue
        if not (bmin <= ask <= bmax): continue
        if liq is not None and liq < minliq: continue
        net = (ref*(1-fee) - ask)/ask*100
        if net < threshold_for(ask, cfg["margin_tiers"]): continue
        st = state.get(iid)
        if st:
            try: last = datetime.datetime.fromisoformat(st["ts"])
            except Exception: last = None
            if last and (now-last).total_seconds() < cooldown and ask >= st.get("ask", 0)*0.95:
                continue
        deals.append({"id":iid,"name":name,"cat":cat,"ask":ask,"med":med,"ref":ref,
                      "net":round(net,1),"profit":round(ref*(1-fee)-ask),
                      "liq":round(liq,1) if liq else "n/a",
                      "block":block_for(ask, cfg["display_blocks"]), "ql":"",
                      "below": sum(1 for a in asks if med is not None and a < med)})
        state[iid] = {"ask": ask, "ts": now.isoformat()}

    print(f"Проверено {checked}, найдено сделок: {len(deals)}")

    # --- список покупок (вишлист): отдельные предметы, которые ты хочешь купить ---
    wish = []
    default_disc = float(cfg.get("wishlist_default_discount_pct", 0) or 0)
    for _wi in load_wishlist():
            iid = str(_wi[0]).strip()
            if not iid or iid.startswith("#"): continue
            name = _wi[1] or iid
            def _f(x):
                try: return float(str(x).strip().replace(" ", "").replace(",", ".")) if x not in (None, "") else None
                except ValueError: return None
            maxp = _f(_wi[2]); disc = _f(_wi[3] if len(_wi) > 3 else None)
            try:
                ask = cheapest_ask(api(f"auction/{iid}/lots?limit={LOTS_LIMIT}", H)); time.sleep(SLEEP_SEC)
                med, _ = market(api(f"auction/{iid}/history?limit={HIST_LIMIT}", H)); time.sleep(SLEEP_SEC)
            except urllib.error.HTTPError as e:
                if e.code == 429: time.sleep(30)
                continue
            except Exception:
                continue
            if ask is None: continue
            if maxp is not None:
                target, rule = maxp, f"цель ≤{sp(maxp)}"
            elif disc is not None and med is not None:
                target, rule = med*(1-disc/100.0), f"−{disc:g}% от медианы"
            elif default_disc and med is not None:
                target, rule = med*(1-default_disc/100.0), f"−{default_disc:g}% от медианы"
            elif med is not None:
                target, rule = med, "ниже медианы"
            else:
                continue
            if ask > target: continue
            key = "wish:" + iid; st = state.get(key)
            if st:
                try: last = datetime.datetime.fromisoformat(st["ts"])
                except Exception: last = None
                if last and (now-last).total_seconds() < cooldown and ask >= st.get("ask", 0)*0.97:
                    continue
            wish.append({"id":iid,"name":name,"ask":round(ask),
                         "med":(round(med) if med else None), "rule":rule})
            state[key] = {"ask": ask, "ts": now.isoformat()}
    print(f"Вишлист: совпадений {len(wish)}")

    # лог истории
    new = not os.path.exists(DEALS_LOG)
    with open(DEALS_LOG, "a", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        if new: w.writerow(["timestamp","id","name","category","ask","market_2nd","market_median","net_margin_pct","profit_per_unit","liq_per_day","block"])
        for d in deals:
            w.writerow([now.isoformat(),d["id"],d["name"],d["cat"],round(d["ask"]),
                        round(d["ref"]) if d.get("ref") is not None else "",
                        round(d["med"]) if d.get("med") is not None else "",
                        d["net"],d["profit"],d["liq"],d["block"]])

    for _m in build_messages(deals, wish, cfg, checked):
        send_telegram(_m)
    json.dump(state, open(STATE, "w", encoding="utf-8"), ensure_ascii=False, indent=0)

def build_messages(deals, wish, cfg, checked):
    """Возвращает СПИСОК сообщений: вишлист — отдельно, и по одному на каждый ценовой блок."""
    msgs = []
    fee = int(cfg["auction_fee"]*100)
    # 1) вишлист — первым отдельным сообщением
    if wish:
        lines = ["🛒 <b>STALZONE — из твоего списка покупок</b>"]
        for w in wish:
            med = f" · медиана {sp(w['med'])}" if w.get("med") else ""
            tag = f" ({w['rule']})" if w.get("rule") else ""
            lines.append(f"• <b>{html.escape(str(w['name']))}</b>\n   купить {sp(w['ask'])}{tag}{med}")
        msgs.append("\n".join(lines))
    # 2) сделки — по одному сообщению на ценовой блок
    if deals:
        sort_key = (lambda d: d["net"]) if cfg.get("sort_within_block") == "margin" else (lambda d: d["profit"])
        maxn = int(cfg.get("max_items_per_block", 10))
        for b in cfg["display_blocks"]:
            label = b.get("label", "")
            grp = sorted([d for d in deals if d["block"] == label], key=sort_key, reverse=True)
            if not grp: continue
            block_profit = sum(d["profit"] for d in grp)
            lines = [f"🟢 <b>STALZONE · 💰 {label}</b> — {len(grp)} лот(ов) · суммарно <b>+{sp(block_profit)}</b> (после {fee}%)"]
            for d in grp[:maxn]:
                link = f"https://stalzone.wiki/items/{section(d['cat'])}/{d['id']}"
                ql = f" <i>[{html.escape(str(d['ql']))}]</i>" if d.get("ql") else ""
                ctx = []
                if d.get("med"): ctx.append(f"медиана {sp(d['med'])}")
                if d.get("below"): ctx.append(f"ниже медианы {d['below']} лот.")
                ctxs = (" · " + " · ".join(ctx)) if ctx else ""
                lines.append(
                    f"• <a href=\"{link}\">{html.escape(str(d['name']))}</a>{ql}\n"
                    f"   предложение {sp(d['ask'])} → 2-я цена {sp(d['ref'])} | <b>+{sp(d['profit'])}</b> ({d['net']}%) · ликв {d['liq']}/д{ctxs}"
                )
            if len(grp) > maxn:
                lines.append(f"   …и ещё {len(grp)-maxn} (см. deals_log.csv)")
            msgs.append("\n".join(lines))
    return msgs

if __name__ == "__main__":
    main()
